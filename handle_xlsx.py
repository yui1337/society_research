from openpyxl import Workbook, load_workbook
from utils import valid_xlsx


def parse_xlsx(path: str) -> Workbook:
    """Table Parsing

    Args:
        path (str): path to table

    Returns:
        Workbook: loaded table
    """
    valid_xlsx(path)
    return load_workbook(path)


def create_result_table(input_table: Workbook, group_size: int, group_id: str) -> Workbook:
    """Creates table as described in technical task

    Args:
        input_table (Workbook): Table from Google Forms, cleared from crap
        group_size (int): Number of academical group members
        group_id (str): Group ID, like 6214 or 6201-FIIT

    Returns:
        Workbook: Table as described in technical task
    """
    wb = Workbook()
    def_sheet = wb.active
    wb.remove(def_sheet) # Удаляем дефолтный лист

    for sheet in range(1,7): # Создаём 6 листов
        new_sheet = wb.create_sheet(title=f"Вопрос № {sheet}")
        new_sheet.cell(1, 1).value = group_id

        for j in range(1, group_size + 1): # Создаём таблицу NxN, N - кол-во чел. в группе
            new_sheet.cell(j + 1, 1).value = j
            new_sheet.cell(1, j + 1).value = j

        for row in range(2, group_size + 2): # Заполняем таблицу нулями
            for col in range(2, group_size + 2):
                new_sheet.cell(row, col).value = 0

        for row in range(1, group_size + 1): # Проставляем единицы
            cur_student_number = int(input_table.active.cell(row + 1, 3).value)
            if not 1 <= cur_student_number <= group_size:
                raise ValueError(f"Invalid student number at row {row + 1}")

            value = input_table.active.cell(row + 1, sheet + 3).value # Считываем строку ответов
            response_value = str(value) if value is not(None or 0) else ""
            try:
                numbers = [int(x) for x in response_value.split() if int(x) <= group_size]
                # Разбиваем ответ на отдельные числа
            except ValueError:
                raise ValueError(f"Invalid response value at input table row {row + 1}, question {sheet}")

            for number in numbers:
                if number == 0:
                    # Студент никого не выбрал, оставляем 0, идём дальше
                    continue
                if not (1 <= number <= group_size):
                    raise ValueError(f"Invalid response value '{number}' at row {row + 1}, question {sheet + 3}")
                new_sheet.cell(cur_student_number + 1, number + 1 ).value = 1

    return wb


def save_xlsx(table: Workbook, path: str) -> None:
    """_summary_

    Args:
        table (Workbook): Table
        path (str): Saved path
    """
    valid_xlsx(path, exists=False)
    table.save(path)

